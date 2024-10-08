import os
import csv
import openpyxl
from tkinter import Tk, filedialog, messagebox, Toplevel, Label

# Função para abrir o explorador de ficheiros e selecionar um ficheiro Excel ou CSV
def select_file(title, filetypes):
    Tk().withdraw()  # Esconder a janela principal do Tkinter
    file = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes
    )
    return file

# Função para ler o valor após "Número da Ordem de Trabalho" num ficheiro CSV
def get_value_from_first_file(file_path):
    if file_path.endswith('.csv'):
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=';')  # Especificar o delimitador como ';'
            for row in reader:
                if "Número da Ordem de Trabalho" in row[0]:  # Verificar se a linha contém "Número da Ordem de Trabalho"
                    return row[1].strip().lstrip('0')  # Remover zeros à esquerda
            return None  # Se não encontrar o valor, retornar None
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return str(sheet['B1'].value).lstrip('0')  # Remover zeros à esquerda se for Excel

# Função para procurar o valor no segundo ficheiro e retornar o valor da coluna E correspondente
def find_value_in_second_file(file_path, search_value):
    search_value = search_value.lstrip('0')
    
    if file_path.endswith('.csv'):
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=';')
            for row in reader:
                if row[0].strip().lstrip('0') == search_value:
                    return row[4].strip()  # Retornar o valor da coluna E
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1, max_col=5, values_only=False):
            if str(row[0].value).lstrip('0') == search_value:
                return row[4].value  # Retornar o valor da coluna E

    return None

# Função para renomear o primeiro ficheiro
def rename_first_file(old_file_path, first_value, second_value):
    directory, old_file_name = os.path.split(old_file_path)
    file_name, file_extension = os.path.splitext(old_file_name)
    
    first_value_cleaned = first_value.lstrip('0')
    
    new_file_name = f"{second_value}_{first_value_cleaned}{file_extension}"
    new_file_path = os.path.join(directory, new_file_name)
    
    try:
        os.rename(old_file_path, new_file_path)
        return True
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao renomear o ficheiro: {e}")
        return False

# Função para mostrar a janela de processamento
def show_processing_window():
    processing_window = Toplevel()
    processing_window.title("A processar...")
    processing_window.geometry("250x100")  # Ajustar a geometria da janela

    label = Label(processing_window, text="A processar, por favor aguarde...")
    label.pack(expand=True)
    
    return processing_window

# Função principal
def main():
    Tk().withdraw()  # Esconder a janela principal do Tkinter
    folder = filedialog.askdirectory(title="Selecione a pasta com ficheiros a ser renomeados")
    if not folder:
        messagebox.showwarning("Atenção", "Não foi selecionada nenhuma pasta.")
        return

    second_file = select_file("Selecione o ficheiro de controlo", [("Excel Files", "*.xlsx;*.xls")])
    if not second_file:
        messagebox.showwarning("Atenção", "Não foi selecionado nenhum ficheiro.")
        return

    count_renamed_files = 0
    processing_window = show_processing_window()  # Mostrar a janela de processamento

    # Atualizar a janela de processamento
    processing_window.update()  # Atualizar a janela para garantir que aparece

    for filename in os.listdir(folder):
        if filename.startswith("Detalhe") and (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            first_file_path = os.path.join(folder, filename)

            value_b1 = get_value_from_first_file(first_file_path)
            if not value_b1:
                continue  # Se não encontrar o valor, continua para o próximo ficheiro

            result = find_value_in_second_file(second_file, value_b1)
            if result:
                if rename_first_file(first_file_path, value_b1, result):
                    count_renamed_files += 1

    processing_window.destroy()  # Fechar a janela de processamento após terminar o loop

    if count_renamed_files > 0:
        messagebox.showinfo("Concluído", f"Foram renomeados {count_renamed_files} ficheiros com sucesso.")
    else:
        messagebox.showinfo("Concluído", "Não foram encontrados ficheiros para renomear.")

# Executar o programa
if __name__ == "__main__":
    main()